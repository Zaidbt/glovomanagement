#!/usr/bin/env python3
"""
Convert Excel product file to Glovo menu JSON format

Usage:
    python scripts/convert-excel-to-glovo-menu.py <input.xlsx> <output.json>

Example:
    python scripts/convert-excel-to-glovo-menu.py ../oulfa.xlsx public/glovo-menu.json
"""

import sys
import json
import openpyxl
from collections import defaultdict

def convert_excel_to_glovo_menu(excel_path, output_path):
    """Convert Excel file to Glovo menu JSON format"""

    print(f"📖 Reading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"📋 Headers: {headers}")

    # Map for collections (categories)
    categories_map = defaultdict(list)
    products = []

    total_rows = 0
    active_count = 0
    inactive_count = 0

    print(f"\n🔄 Processing products...")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        data = dict(zip(headers, row))

        # Skip rows without SKU or name
        if not data.get('sku') or not data.get('name'):
            continue

        total_rows += 1

        # Extract data
        sku = str(data['sku']).strip()
        name = str(data['name']).strip()
        price = float(data.get('price', 0))
        active = bool(data.get('active', False))
        category1 = data.get('category 1')
        category2 = data.get('category 2')

        # Track counts
        if active:
            active_count += 1
        else:
            inactive_count += 1

        # Create product object
        product = {
            "id": sku,
            "name": name,
            "description": name,  # Use name as description
            "price": price,
            "image_url": "https://i.imgur.com/Qj5MlEH.png",  # Default image
            "attributes_groups": [],
            "available": active
        }

        products.append(product)

        # Add to categories
        if category1:
            categories_map[category1].append(sku)
        if category2 and category2 != category1:
            categories_map[category2].append(sku)

        # Progress indicator
        if total_rows % 100 == 0:
            print(f"  Processed {total_rows} products...")

    print(f"\n✅ Processed {total_rows} products")
    print(f"  - Active: {active_count}")
    print(f"  - Inactive: {inactive_count}")
    print(f"  - Categories: {len(categories_map)}")

    # Create collections from categories
    collections = []
    for idx, (category_name, product_ids) in enumerate(sorted(categories_map.items())):
        if not product_ids:  # Skip empty categories
            continue

        collection = {
            "name": category_name,
            "position": idx,
            "image_url": "https://i.imgur.com/Qj5MlEH.png",
            "sections": [
                {
                    "name": category_name,
                    "position": 1,
                    "products": product_ids
                }
            ]
        }
        collections.append(collection)

    # Create final menu structure
    menu = {
        "attributes": [],  # Can add later if needed
        "attribute_groups": [],  # Can add later if needed
        "products": products,
        "collections": collections,
        "supercollections": [
            {
                "name": "Tous les produits",
                "position": 0,
                "image_url": "https://i.imgur.com/Qj5DlEH.png",
                "collections": [col["name"] for col in collections]
            }
        ] if collections else []
    }

    # Write to JSON file
    print(f"\n💾 Writing to: {output_path}")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(menu, f, ensure_ascii=False, indent=2)

    print(f"\n🎉 Conversion complete!")
    print(f"  - Total products: {len(products)}")
    print(f"  - Active products: {active_count}")
    print(f"  - Collections: {len(collections)}")
    print(f"  - Output file: {output_path}")

    # Show sample of first product
    if products:
        print(f"\n📦 Sample product:")
        print(json.dumps(products[0], ensure_ascii=False, indent=2))

    return menu

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python convert-excel-to-glovo-menu.py <input.xlsx> <output.json>")
        print("\nExample:")
        print("  python scripts/convert-excel-to-glovo-menu.py ../oulfa.xlsx public/glovo-menu.json")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    try:
        convert_excel_to_glovo_menu(input_file, output_file)
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
